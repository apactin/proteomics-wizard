#!/usr/bin/env python3
"""
splitting.py

Skyline split-peak caller with Peak Boundary guidance.

Wizard-ready refactor:
- No work at import time
- Parameterized IO via run(...)
- Fixes sheet-name bug for consistent hit sheet
- Safer propagation keys (Site + Uniprot ID if available)
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font


LogFn = Callable[[str], None]


# -----------------------------
# Fixed Skyline schema (chrom traces TSV)
# -----------------------------

SKY_COL = {
    "filename": "FileName",
    "peptide_mod_seq": "PeptideModifiedSequence",
    "precursor_charge": "PrecursorCharge",
    "product_mz": "ProductMz",
    "fragment_ion": "FragmentIon",
    "product_charge": "ProductCharge",
    "isotope_label": "IsotopeLabelType",
    "total_area": "TotalArea",
    "times": "Times",
    "intensities": "Intensities",
}

REQUIRED_SKY_COLS = list(SKY_COL.values())


# -----------------------------
# Config / thresholds
# -----------------------------

@dataclass
class SplitConfig:
    min_peak_frac_of_global_max: float = 0.10
    min_peak2_frac_of_peak1: float = 0.20
    smooth_window_points: int = 3
    complete_valley_frac_max: float = 0.35
    partial_valley_frac_max: float = 0.75
    min_rt_separation: float = 0.05
    consensus_rule: str = "majority_then_worst"
    max_rt_delta_for_split: float = 1.0
    boundary_pad_min: float = 0.15

    shoulder_valley_frac_max: float = 0.90
    shoulder_peak2_ratio_min: float = 0.40
    shoulder_rt_delta_max: float = 0.30
    shoulder_drop_frac_min: float = 0.15
    shoulder_recover_ratio_min: float = 0.60


SPLIT_ORDER = {"Single": 0, "Partial": 1, "Complete": 2}


# -----------------------------
# IO models
# -----------------------------

@dataclass(frozen=True)
class SplittingInputs:
    sites_xlsx: Path
    sky_ms2_tsv: Path
    sky_ms3_tsv: Path
    ms2_peaks_csv: Path
    ms3_peaks_csv: Path


@dataclass(frozen=True)
class SplittingOutputs:
    output_xlsx: Path
    traces_parquet: Path
    n_unique_sites: int
    n_chrom_rows: int
    n_trace_rows: int


def _log(msg: str, log: Optional[LogFn]) -> None:
    if log:
        log(msg)
    else:
        print(msg)


# -----------------------------
# Peptide normalization
# -----------------------------

_charge_suffix_re = re.compile(r"\++\s*$")
_bracket_re = re.compile(r"\[([^\]]+)\]")
_num_re = re.compile(r"^[+-]?\d+(\.\d+)?$")


def _format_mass_token(tok: str) -> str:
    t = str(tok).strip()
    if not _num_re.match(t):
        return t
    val = float(t)
    sign = "+" if val >= 0 else "-"
    aval = abs(val)
    if aval.is_integer():
        core = str(int(aval))
    else:
        core = f"{aval:.10f}".rstrip("0").rstrip(".")
    return f"{sign}{core}"


def normalize_one_modified_peptide(seq: str) -> str:
    if seq is None:
        return ""
    s = str(seq).strip()
    if not s or s.lower() == "nan":
        return ""
    s = re.sub(r"\s+", "", s)
    s = _charge_suffix_re.sub("", s)

    def repl(m: re.Match) -> str:
        return "[" + _format_mass_token(m.group(1)) + "]"

    return _bracket_re.sub(repl, s)


def split_and_normalize_variants(cell: object) -> List[str]:
    if cell is None:
        return []
    raw = str(cell).strip()
    if not raw or raw.lower() == "nan":
        return []
    parts = [p.strip() for p in raw.split(";")]
    seen, out = set(), []
    for p in parts:
        n = normalize_one_modified_peptide(p)
        if n and n not in seen:
            seen.add(n)
            out.append(n)
    return out


# -----------------------------
# Parse Times/Intensities
# -----------------------------

_NUM_SPLIT_RE = re.compile(r"[,\s;]+")


def parse_num_list(s: object) -> np.ndarray:
    if s is None:
        return np.array([], dtype=float)
    if isinstance(s, float) and np.isnan(s):
        return np.array([], dtype=float)
    txt = str(s).strip()
    if not txt:
        return np.array([], dtype=float)
    parts = [p for p in _NUM_SPLIT_RE.split(txt) if p != ""]
    out = []
    for p in parts:
        try:
            out.append(float(p))
        except ValueError:
            continue
    return np.asarray(out, dtype=float)


def validate_skyline_schema(df: pd.DataFrame, label: str) -> None:
    missing = [c for c in REQUIRED_SKY_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f"{label} Skyline TSV is missing required columns: {missing}\n"
            f"Available columns: {list(df.columns)}"
        )


def load_skyline_tsv(path: Path, label: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep="\t", dtype=str)
    validate_skyline_schema(df, label)
    for c in ["PrecursorCharge", "ProductMz", "ProductCharge", "TotalArea"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.copy()
    df["sky_norm_peptide"] = df[SKY_COL["peptide_mod_seq"]].apply(normalize_one_modified_peptide)

    # only precursor traces (much faster downstream)
    frag = df[SKY_COL["fragment_ion"]].astype(str).str.lower().str.strip()
    df = df[frag == "precursor"].copy()
    return df


# -----------------------------
# Load Peak Boundaries CSV
# -----------------------------

PEAKS_REQUIRED = [
    "File Name",
    "Peptide Modified Sequence",
    "Min Start Time",
    "Max End Time",
    "Precursor Charge",
    "PrecursorIsDecoy",
]


def load_peaks_report_csv(path: Path, label: str) -> Dict[Tuple[str, str, int], Tuple[float, float]]:
    """
    Returns dict:
      (FileName, norm_peptide, precursor_charge) -> (min_start_time, max_end_time)

    - Drops decoys.
    - If multiple rows map to same key, keeps the broadest window.
    """
    if not path.exists():
        raise FileNotFoundError(f"{label} peaks report not found: {path}")

    df = pd.read_csv(path, dtype=str)
    missing = [c for c in PEAKS_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(
            f"{label} peaks report CSV missing required columns: {missing}\n"
            f"Available columns: {list(df.columns)}"
        )

    def is_decoy(x: object) -> bool:
        s = str(x).strip().lower()
        return s in {"true", "t", "1", "yes", "y"}

    df = df.copy()
    df = df[~df["PrecursorIsDecoy"].apply(is_decoy)].copy()

    df["Min Start Time"] = pd.to_numeric(df["Min Start Time"], errors="coerce")
    df["Max End Time"] = pd.to_numeric(df["Max End Time"], errors="coerce")
    df["Precursor Charge"] = pd.to_numeric(df["Precursor Charge"], errors="coerce")

    df["norm_pep"] = df["Peptide Modified Sequence"].apply(normalize_one_modified_peptide)

    out: Dict[Tuple[str, str, int], Tuple[float, float]] = {}
    for _, r in df.iterrows():
        fn = str(r["File Name"]).strip()
        pep = str(r["norm_pep"]).strip()
        chg = r["Precursor Charge"]
        st = r["Min Start Time"]
        en = r["Max End Time"]

        if not fn or not pep or not np.isfinite(chg) or not np.isfinite(st) or not np.isfinite(en):
            continue
        chg_i = int(chg)
        st_f = float(st)
        en_f = float(en)
        if en_f <= st_f:
            continue

        key = (fn, pep, chg_i)
        if key not in out:
            out[key] = (st_f, en_f)
        else:
            prev_st, prev_en = out[key]
            out[key] = (min(prev_st, st_f), max(prev_en, en_f))

    return out


# -----------------------------
# Peak analysis
# -----------------------------

def moving_average(y: np.ndarray, w: int) -> np.ndarray:
    if w is None or w <= 1:
        return y
    w = int(w)
    pad = w // 2
    ypad = np.pad(y, (pad, pad), mode="edge")
    kernel = np.ones(w, dtype=float) / float(w)
    return np.convolve(ypad, kernel, mode="valid")


def local_maxima_indices(y: np.ndarray) -> np.ndarray:
    if len(y) < 3:
        return np.array([], dtype=int)
    dy1 = y[1:-1] - y[:-2]
    dy2 = y[1:-1] - y[2:]
    mask = (dy1 > 0) & (dy2 >= 0)
    return (np.where(mask)[0] + 1).astype(int)


def valley_between(rt: np.ndarray, y: np.ndarray, i1: int, i2: int) -> Tuple[float, float, int]:
    lo, hi = (i1, i2) if i1 < i2 else (i2, i1)
    seg = y[lo:hi + 1]
    j = int(np.argmin(seg))
    valley_idx = lo + j
    return float(rt[valley_idx]), float(y[valley_idx]), valley_idx


@dataclass
class ChromatogramResult:
    call: str
    peak_count_considered: int
    peak1_rt: Optional[float]
    peak2_rt: Optional[float]
    peak1_height: Optional[float]
    peak2_height: Optional[float]
    peak2_ratio: Optional[float]
    rt_delta: Optional[float]
    valley_rt: Optional[float]
    valley_intensity: Optional[float]
    valley_frac: Optional[float]
    split_strength: Optional[float]
    debug: Dict


def _pick_peak1_index(
    rt: np.ndarray,
    y_smooth: np.ndarray,
    peaks: np.ndarray,
    cfg: SplitConfig,
    boundary_start: Optional[float],
    boundary_end: Optional[float],
) -> Tuple[int, Dict]:
    dbg: Dict = {}
    global_idx = int(np.argmax(y_smooth))
    if boundary_start is None or boundary_end is None:
        dbg["peak1_pick"] = "global_max_no_boundary"
        return global_idx, dbg

    if not np.isfinite(boundary_start) or not np.isfinite(boundary_end) or boundary_end <= boundary_start:
        dbg["peak1_pick"] = "global_max_bad_boundary"
        dbg["boundary_start"] = boundary_start
        dbg["boundary_end"] = boundary_end
        return global_idx, dbg

    pad = float(cfg.boundary_pad_min)
    lo = float(boundary_start) - pad
    hi = float(boundary_end) + pad
    dbg["boundary_start"] = float(boundary_start)
    dbg["boundary_end"] = float(boundary_end)
    dbg["boundary_pad_min"] = pad
    dbg["boundary_window_lo"] = lo
    dbg["boundary_window_hi"] = hi

    in_win = peaks[(rt[peaks] >= lo) & (rt[peaks] <= hi)]
    if in_win.size == 0:
        dbg["peak1_pick"] = "global_max_no_peaks_in_boundary_window"
        return global_idx, dbg

    peak1_idx = int(in_win[np.argmax(y_smooth[in_win])])
    dbg["peak1_pick"] = "max_in_boundary_window"
    return peak1_idx, dbg


def analyze_chromatogram(
    rt: np.ndarray,
    intensity: np.ndarray,
    cfg: SplitConfig,
    boundary_start: Optional[float] = None,
    boundary_end: Optional[float] = None,
) -> ChromatogramResult:
    m = np.isfinite(rt) & np.isfinite(intensity)
    rt = rt[m]
    intensity = intensity[m]
    if rt.size < 5:
        return ChromatogramResult(
            call="Single", peak_count_considered=0,
            peak1_rt=None, peak2_rt=None,
            peak1_height=None, peak2_height=None,
            peak2_ratio=None, rt_delta=None,
            valley_rt=None, valley_intensity=None, valley_frac=None,
            split_strength=None,
            debug={"reason": "too_few_points", "n": int(rt.size)},
        )

    order = np.argsort(rt)
    rt = rt[order]
    y_raw = intensity[order].astype(float)

    y_smooth = moving_average(y_raw, cfg.smooth_window_points)
    gmax = float(np.nanmax(y_smooth)) if y_smooth.size else float("nan")
    if not np.isfinite(gmax) or gmax <= 0:
        return ChromatogramResult(
            call="Single", peak_count_considered=0,
            peak1_rt=None, peak2_rt=None,
            peak1_height=None, peak2_height=None,
            peak2_ratio=None, rt_delta=None,
            valley_rt=None, valley_intensity=None, valley_frac=None,
            split_strength=None,
            debug={"reason": "nonpositive_max", "gmax_smooth": gmax},
        )

    peaks = local_maxima_indices(y_smooth)
    global_idx = int(np.argmax(y_smooth))
    if global_idx not in set(peaks.tolist()):
        peaks = np.unique(np.append(peaks, global_idx))

    keep1 = peaks[y_smooth[peaks] >= cfg.min_peak_frac_of_global_max * gmax]
    if keep1.size == 0:
        keep1 = np.array([global_idx], dtype=int)

    peak1_idx, peak1_dbg = _pick_peak1_index(
        rt=rt,
        y_smooth=y_smooth,
        peaks=keep1,
        cfg=cfg,
        boundary_start=boundary_start,
        boundary_end=boundary_end,
    )

    peak1_rt = float(rt[peak1_idx])
    peak1_h_raw = float(y_raw[peak1_idx])

    cand = []
    for i in keep1:
        i = int(i)
        if i == peak1_idx:
            continue
        if y_smooth[i] < cfg.min_peak2_frac_of_peak1 * y_smooth[peak1_idx]:
            continue
        if abs(float(rt[i]) - peak1_rt) < cfg.min_rt_separation:
            continue
        cand.append(i)
    cand = np.array(sorted(set(cand)), dtype=int)

    # shoulder-only fallback if no local maxima candidate exists
    if cand.size == 0:
        shoulder_dbg = {"shoulder_fallback": True}
        is_shoulder = False
        dip_idx = None
        rec_idx = None

        win_lo = peak1_rt
        win_hi = peak1_rt + float(cfg.shoulder_rt_delta_max)
        shoulder_dbg["shoulder_window_lo"] = float(win_lo)
        shoulder_dbg["shoulder_window_hi"] = float(win_hi)

        idx_win = np.where((rt >= win_lo) & (rt <= win_hi))[0]
        if idx_win.size >= 4:
            y_w = y_raw[idx_win]

            dip_rel = int(np.argmin(y_w[1:])) + 1
            dip_idx = int(idx_win[dip_rel])
            dip_y = float(y_raw[dip_idx])
            dip_rt = float(rt[dip_idx])

            if dip_rel < (y_w.size - 1):
                rec_rel = int(np.argmax(y_w[dip_rel:])) + dip_rel
                rec_idx = int(idx_win[rec_rel])
                rec_y = float(y_raw[rec_idx])
                rec_rt = float(rt[rec_idx])

                drop_frac = (peak1_h_raw - dip_y) / peak1_h_raw if peak1_h_raw > 0 else 0.0
                recover_ratio = rec_y / peak1_h_raw if peak1_h_raw > 0 else 0.0
                rt_delta2 = abs(rec_rt - peak1_rt)

                shoulder_dbg.update({
                    "dip_rt": dip_rt,
                    "dip_y": dip_y,
                    "recover_rt": rec_rt,
                    "recover_y": rec_y,
                    "drop_frac": float(drop_frac),
                    "recover_ratio": float(recover_ratio),
                    "recover_rt_delta": float(rt_delta2),
                })

                if (drop_frac >= cfg.shoulder_drop_frac_min and
                    recover_ratio >= cfg.shoulder_recover_ratio_min and
                    rt_delta2 <= cfg.shoulder_rt_delta_max):
                    is_shoulder = True
        else:
            shoulder_dbg["reason"] = "too_few_points_in_shoulder_window"

        if is_shoulder and rec_idx is not None and dip_idx is not None:
            peak2_h = float(y_raw[rec_idx])
            denom = float(min(peak1_h_raw, peak2_h))
            valley_frac_val = (float(y_raw[dip_idx]) / denom) if denom > 0 else None

            return ChromatogramResult(
                call="Partial",
                peak_count_considered=int(keep1.size),
                peak1_rt=peak1_rt,
                peak2_rt=float(rt[rec_idx]),
                peak1_height=peak1_h_raw,
                peak2_height=peak2_h,
                peak2_ratio=(peak2_h / peak1_h_raw) if peak1_h_raw > 0 else None,
                rt_delta=float(abs(float(rt[rec_idx]) - peak1_rt)),
                valley_rt=float(rt[dip_idx]),
                valley_intensity=float(y_raw[dip_idx]),
                valley_frac=valley_frac_val,
                split_strength=None,
                debug={
                    "gmax_smooth": gmax,
                    "smooth_window_points": cfg.smooth_window_points,
                    "n_maxima": int(peaks.size),
                    "n_after_10pct": int(keep1.size),
                    "n_peak2_candidates": 0,
                    "note": "partial_due_to_shoulder_fallback_no_local_max_in_smooth",
                    "boundary_start": boundary_start,
                    "boundary_end": boundary_end,
                    **peak1_dbg,
                    **shoulder_dbg,
                },
            )

        return ChromatogramResult(
            call="Single", peak_count_considered=int(keep1.size),
            peak1_rt=peak1_rt, peak2_rt=None,
            peak1_height=peak1_h_raw, peak2_height=None,
            peak2_ratio=None, rt_delta=None,
            valley_rt=None, valley_intensity=None, valley_frac=None,
            split_strength=None,
            debug={
                "gmax_smooth": gmax,
                "smooth_window_points": cfg.smooth_window_points,
                "n_maxima": int(peaks.size),
                "n_after_10pct": int(keep1.size),
                "n_peak2_candidates": 0,
                "note": "peak1_boundary_anchored_peak2_global",
                "boundary_start": boundary_start,
                "boundary_end": boundary_end,
                **peak1_dbg,
                **shoulder_dbg,
            },
        )

    # choose best peak2 among candidates
    best = None
    for p2 in cand:
        peak2_rt = float(rt[p2])
        peak2_h_raw = float(y_raw[p2])
        v_rt, v_int_raw, _ = valley_between(rt, y_raw, peak1_idx, p2)

        denom = min(peak1_h_raw, peak2_h_raw)
        v_frac = float(v_int_raw / denom) if denom > 0 else 1.0
        rt_delta = abs(peak2_rt - peak1_rt)
        split_strength = float(1.0 - v_frac)

        score = (v_frac, -peak2_h_raw)
        if best is None or score < best["score"]:
            best = {
                "score": score,
                "peak2_h_raw": peak2_h_raw,
                "peak2_rt": peak2_rt,
                "valley_rt": v_rt,
                "valley_int_raw": v_int_raw,
                "valley_frac": v_frac,
                "rt_delta": rt_delta,
                "split_strength": split_strength,
            }

    assert best is not None

    if best["rt_delta"] is not None and best["rt_delta"] > cfg.max_rt_delta_for_split:
        return ChromatogramResult(
            call="Single",
            peak_count_considered=int(keep1.size),
            peak1_rt=peak1_rt,
            peak2_rt=None,
            peak1_height=peak1_h_raw,
            peak2_height=None,
            peak2_ratio=None,
            rt_delta=float(best["rt_delta"]),
            valley_rt=None,
            valley_intensity=None,
            valley_frac=None,
            split_strength=None,
            debug={
                "gmax_smooth": gmax,
                "smooth_window_points": cfg.smooth_window_points,
                "n_maxima": int(peaks.size),
                "n_after_10pct": int(keep1.size),
                "n_peak2_candidates": int(cand.size),
                "note": "forced_single_due_to_large_rt_delta",
                "rt_delta": float(best["rt_delta"]),
                "max_rt_delta_for_split": float(cfg.max_rt_delta_for_split),
                "boundary_start": boundary_start,
                "boundary_end": boundary_end,
                **peak1_dbg,
            },
        )

    if best["valley_frac"] <= cfg.complete_valley_frac_max:
        call = "Complete"
    elif best["valley_frac"] <= cfg.partial_valley_frac_max:
        call = "Partial"
    else:
        call = "Single"

    peak2_ratio_out = (float(best["peak2_h_raw"]) / peak1_h_raw) if peak1_h_raw > 0 else None

    if call == "Single" and peak2_ratio_out is not None:
        is_shoulder = (
            (peak2_ratio_out >= cfg.shoulder_peak2_ratio_min) and
            (float(best["rt_delta"]) <= cfg.shoulder_rt_delta_max) and
            (float(best["valley_frac"]) <= cfg.shoulder_valley_frac_max)
        )
        if is_shoulder:
            call = "Partial"

    return ChromatogramResult(
        call=call,
        peak_count_considered=int(keep1.size),
        peak1_rt=peak1_rt,
        peak2_rt=float(best["peak2_rt"]),
        peak1_height=peak1_h_raw,
        peak2_height=float(best["peak2_h_raw"]),
        peak2_ratio=peak2_ratio_out,
        rt_delta=float(best["rt_delta"]),
        valley_rt=float(best["valley_rt"]),
        valley_intensity=float(best["valley_int_raw"]),
        valley_frac=float(best["valley_frac"]),
        split_strength=float(best["split_strength"]),
        debug={
            "gmax_smooth": gmax,
            "smooth_window_points": cfg.smooth_window_points,
            "n_maxima": int(peaks.size),
            "n_after_10pct": int(keep1.size),
            "n_peak2_candidates": int(cand.size),
            "note": "peak1_boundary_anchored_peak2_global",
            "max_rt_delta_for_split": float(cfg.max_rt_delta_for_split),
            "boundary_start": boundary_start,
            "boundary_end": boundary_end,
            **peak1_dbg,
        },
    )


def consensus_call(calls: List[str], cfg: SplitConfig) -> str:
    calls = [c for c in calls if c in SPLIT_ORDER]
    if not calls:
        return "Single"
    if cfg.consensus_rule == "worst":
        return max(calls, key=lambda c: SPLIT_ORDER[c])
    counts: Dict[str, int] = {}
    for c in calls:
        counts[c] = counts.get(c, 0) + 1
    best_count = max(counts.values())
    winners = [c for c, n in counts.items() if n == best_count]
    if len(winners) == 1:
        return winners[0]
    return max(winners, key=lambda c: SPLIT_ORDER[c])


# -----------------------------
# Excel helpers
# -----------------------------

def add_chromatograms_sheet(xlsx_path: Path, chrom_rows: List[Dict], sheet_name: str = "Chromatograms") -> None:
    wb = load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    cols = [
        "trace_id", "ms_level", "Gene Name", "Site", "Uniprot ID",
        "Matching modified peptides", "PeptideVariant",
        "FileName", "PrecursorCharge", "ProductMz", "ProductCharge",
        "IsotopeLabelType", "TotalArea",
        "call", "peak1_rt", "peak2_rt", "peak2_ratio",
        "valley_rt", "valley_frac", "split_strength",
        "boundary_start_time", "boundary_end_time", "rt_raw", "intensity_raw",
    ]

    for j, c in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=c).font = Font(bold=True)

    for i, r in enumerate(chrom_rows, start=2):
        for j, c in enumerate(cols, start=1):
            v = r.get(c, "")
            ws.cell(row=i, column=j, value="" if (isinstance(v, float) and np.isnan(v)) else v)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(cols)).column_letter}1"
    wb.save(xlsx_path)


# -----------------------------
# Trace ID + Parquet writing
# -----------------------------

def make_trace_id(ms_level: str, site: str, pep_variant: str, file_name: str, product_mz: object, precursor_charge: object, isotope: str) -> str:
    import hashlib
    base = "|".join([
        str(ms_level or ""),
        str(site or ""),
        str(pep_variant or ""),
        str(file_name or ""),
        str(product_mz if product_mz is not None else ""),
        str(precursor_charge if precursor_charge is not None else ""),
        str(isotope or ""),
    ])
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]


def write_traces_parquet(rows: List[Dict], out_path: Path) -> None:
    try:
        import pyarrow  # noqa: F401
    except Exception as e:
        raise RuntimeError(
            "Writing parquet requires pyarrow. Install with: pip install pyarrow\n"
            f"Original error: {e}"
        )

    df = pd.DataFrame(rows)
    df["rt"] = df["rt"].apply(lambda x: list(map(float, x)) if x is not None else [])
    df["intensity"] = df["intensity"].apply(lambda x: list(map(float, x)) if x is not None else [])
    df.to_parquet(out_path, index=False)


# -----------------------------
# Workbook propagation
# -----------------------------

def propagate_to_sheets(
    xlsx_in: Path,
    xlsx_out: Path,
    unique_sites_updates: pd.DataFrame,
    *,
    key_cols: List[str],
) -> None:
    xl = pd.ExcelFile(xlsx_in)
    sheets = {name: xl.parse(name) for name in xl.sheet_names}
    if "Unique_Sites" not in sheets:
        raise ValueError("Input workbook missing sheet 'Unique_Sites'.")
    sheets["Unique_Sites"] = unique_sites_updates

    # ✅ FIXED: correct consistent sheet name
    target_sheets = ["Hits_MS2", "Hits_MS3", "Hits_Consistent_MS2_MS3"]

    for sname in target_sheets:
        if sname not in sheets:
            continue
        base = sheets[sname].copy()

        common_keys = [c for c in key_cols if c in base.columns and c in unique_sites_updates.columns]
        if not common_keys:
            continue

        # Use a unique lookup to avoid row-multiplication
        take_cols = list(dict.fromkeys(common_keys + [c for c in unique_sites_updates.columns if c.startswith("sky_")]))
        lookup = unique_sites_updates[take_cols].drop_duplicates(subset=common_keys, keep="first")

        sheets[sname] = base.merge(lookup, on=common_keys, how="left")

    with pd.ExcelWriter(xlsx_out, engine="openpyxl") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name, index=False)


# -----------------------------
# Skyline analysis for peptide variants
# -----------------------------

def analyze_skyline_for_peptide_variants(
    skyline_df: pd.DataFrame,
    peptide_variants_norm: List[str],
    cfg: SplitConfig,
    peaks_lookup: Dict[Tuple[str, str, int], Tuple[float, float]],
):
    best_cons, best_per, best_variant, best_n = "Single", [], "", 0

    for variant in peptide_variants_norm:
        dfv = skyline_df[skyline_df["sky_norm_peptide"] == variant]
        if dfv.empty:
            continue

        per, calls = [], []
        for _, r in dfv.iterrows():
            rt = parse_num_list(r[SKY_COL["times"]])
            y = parse_num_list(r[SKY_COL["intensities"]])
            n = min(rt.size, y.size)
            rt, y = rt[:n], y[:n]

            fn = str(r.get(SKY_COL["filename"], "") or "").strip()
            chg = r.get(SKY_COL["precursor_charge"], None)
            chg_i: Optional[int] = None
            if chg is not None and np.isfinite(chg):
                try:
                    chg_i = int(chg)
                except Exception:
                    chg_i = None

            bstart = bend = None
            if fn and chg_i is not None:
                key = (fn, variant, chg_i)
                if key in peaks_lookup:
                    bstart, bend = peaks_lookup[key]

            res = analyze_chromatogram(rt, y, cfg, boundary_start=bstart, boundary_end=bend)
            calls.append(res.call)

            per.append({
                "PeptideVariant": variant,
                "FileName": r.get(SKY_COL["filename"], None),
                "PrecursorCharge": r.get(SKY_COL["precursor_charge"], None),
                "IsotopeLabelType": r.get(SKY_COL["isotope_label"], None),
                "ProductMz": r.get(SKY_COL["product_mz"], None),
                "ProductCharge": r.get(SKY_COL["product_charge"], None),
                "TotalArea": r.get(SKY_COL["total_area"], None),
                "boundary_start_time": bstart,
                "boundary_end_time": bend,
                "rt": rt,
                "intensity": y,
                **asdict(res),
            })

        cons = consensus_call(calls, cfg)
        ntr = len(per)
        if ntr > best_n:
            best_n, best_cons, best_per, best_variant = ntr, cons, per, variant

    return best_cons, best_per, best_variant, best_n


def pick_representative(per_list: List[Dict], consensus: str) -> Dict:
    if not per_list:
        return {}
    subset = [p for p in per_list if p.get("call") == consensus] or per_list
    subset_sorted = sorted(
        subset,
        key=lambda p: (
            SPLIT_ORDER.get(p.get("call", "Single"), 0),
            (p.get("split_strength") if p.get("split_strength") is not None else -1e9),
        ),
        reverse=True,
    )
    return subset_sorted[0] if subset_sorted else {}


# -----------------------------
# Public run()
# -----------------------------

def run(
    inputs: SplittingInputs,
    out_dir: Path,
    *,
    cfg: Optional[SplitConfig] = None,
    output_xlsx_name: Optional[str] = None,
    traces_parquet_name: str = "chrom_traces.parquet",
    log: Optional[LogFn] = None,
) -> SplittingOutputs:
    cfg = cfg or SplitConfig()
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # validate inputs exist
    for p in [inputs.sites_xlsx, inputs.sky_ms2_tsv, inputs.sky_ms3_tsv, inputs.ms2_peaks_csv, inputs.ms3_peaks_csv]:
        if not Path(p).exists():
            raise FileNotFoundError(f"Missing required input: {p}")

    ms2_peaks = load_peaks_report_csv(inputs.ms2_peaks_csv, "MS2")
    ms3_peaks = load_peaks_report_csv(inputs.ms3_peaks_csv, "MS3")

    sites_xl = pd.ExcelFile(inputs.sites_xlsx)
    unique = sites_xl.parse("Unique_Sites")
    if "Matching modified peptides" not in unique.columns:
        raise ValueError("Unique_Sites missing column 'Matching modified peptides'.")

    ms2_df = load_skyline_tsv(inputs.sky_ms2_tsv, "MS2")
    ms3_df = load_skyline_tsv(inputs.sky_ms3_tsv, "MS3")

    col_gene = "Gene Name" if "Gene Name" in unique.columns else None
    col_site = "Site" if "Site" in unique.columns else None
    col_up = "Uniprot ID" if "Uniprot ID" in unique.columns else None

    # choose join keys for propagation (safer than using modified peptide text)
    key_cols: List[str] = []
    if col_site:
        key_cols.append("Site")
    if col_up:
        key_cols.append("Uniprot ID")

    _log(f"[splitting] Join keys for propagation: {key_cols if key_cols else '(none)'}", log)

    sky_norm_all, sky_norm_chosen = [], []
    ms2_calls, ms3_calls = [], []
    ms2_ntraces, ms3_ntraces = [], []
    ms2_detail, ms3_detail = [], []

    ms2_peak1_rt, ms2_peak2_rt, ms2_peak2_ratio, ms2_valley_frac, ms2_split_strength = [], [], [], [], []
    ms3_peak1_rt, ms3_peak2_rt, ms3_peak2_ratio, ms3_valley_frac, ms3_split_strength = [], [], [], [], []

    chrom_rows_excel: List[Dict] = []
    trace_rows_parquet: List[Dict] = []

    def _vec_to_excel_text(a: np.ndarray) -> str:
        if a is None:
            return "[]"
        a = np.asarray(a, dtype=float)
        if a.size == 0:
            return "[]"
        if np.nanmax(np.abs(a)) < 1e4:
            a2 = np.round(a, 6)
        else:
            a2 = np.round(a, 4)
        return json.dumps([float(x) for x in a2.tolist()], separators=(",", ":"))

    def emit_traces(per_list: List[Dict], ms_level: str, gene: str, site: str, unip: str, match_cell: str) -> None:
        for d in per_list:
            rt = np.asarray(d.get("rt", []), dtype=float)
            y = np.asarray(d.get("intensity", []), dtype=float)
            if rt.size == 0 or y.size == 0:
                continue

            trace_id = make_trace_id(
                ms_level=ms_level,
                site=str(site),
                pep_variant=str(d.get("PeptideVariant", "")),
                file_name=str(d.get("FileName", "")),
                product_mz=d.get("ProductMz", ""),
                precursor_charge=d.get("PrecursorCharge", ""),
                isotope=str(d.get("IsotopeLabelType", "")),
            )

            chrom_rows_excel.append({
                "trace_id": trace_id,
                "ms_level": ms_level,
                "Gene Name": gene,
                "Site": site,
                "Uniprot ID": unip,
                "Matching modified peptides": match_cell,
                "PeptideVariant": d.get("PeptideVariant"),
                "FileName": d.get("FileName"),
                "PrecursorCharge": d.get("PrecursorCharge"),
                "ProductMz": d.get("ProductMz"),
                "ProductCharge": d.get("ProductCharge"),
                "IsotopeLabelType": d.get("IsotopeLabelType"),
                "TotalArea": d.get("TotalArea"),
                "call": d.get("call"),
                "peak1_rt": d.get("peak1_rt"),
                "peak2_rt": d.get("peak2_rt"),
                "peak2_ratio": d.get("peak2_ratio"),
                "valley_rt": d.get("valley_rt"),
                "valley_frac": d.get("valley_frac"),
                "split_strength": d.get("split_strength"),
                "boundary_start_time": d.get("boundary_start_time"),
                "boundary_end_time": d.get("boundary_end_time"),
                "rt_raw": _vec_to_excel_text(rt),
                "intensity_raw": _vec_to_excel_text(y),
            })

            trace_rows_parquet.append({
                "trace_id": trace_id,
                "ms_level": ms_level,
                "Gene Name": gene,
                "Site": site,
                "Uniprot ID": unip,
                "Matching modified peptides": match_cell,
                "PeptideVariant": d.get("PeptideVariant"),
                "FileName": d.get("FileName"),
                "PrecursorCharge": d.get("PrecursorCharge"),
                "ProductMz": d.get("ProductMz"),
                "ProductCharge": d.get("ProductCharge"),
                "IsotopeLabelType": d.get("IsotopeLabelType"),
                "TotalArea": d.get("TotalArea"),
                "call": d.get("call"),
                "peak1_rt": d.get("peak1_rt"),
                "peak2_rt": d.get("peak2_rt"),
                "peak2_ratio": d.get("peak2_ratio"),
                "valley_rt": d.get("valley_rt"),
                "valley_frac": d.get("valley_frac"),
                "split_strength": d.get("split_strength"),
                "boundary_start_time": d.get("boundary_start_time"),
                "boundary_end_time": d.get("boundary_end_time"),
                "rt": rt.tolist(),
                "intensity": y.tolist(),
            })

    n_rows = len(unique)
    for idx, row in unique.iterrows():
        if (idx % 250) == 0:
            _log(f"[splitting] Processing {idx+1}/{n_rows} sites...", log)

        variants = split_and_normalize_variants(row.get("Matching modified peptides", ""))
        sky_norm_all.append(";".join(variants))

        gene = str(row[col_gene]) if col_gene else ""
        site = str(row[col_site]) if col_site else ""
        unip = str(row[col_up]) if col_up else ""
        match_cell = row.get("Matching modified peptides", "")

        if not variants:
            sky_norm_chosen.append("")
            ms2_calls.append("Single")
            ms3_calls.append("Single")
            ms2_ntraces.append(0)
            ms3_ntraces.append(0)
            ms2_detail.append("[]")
            ms3_detail.append("[]")
            for lst in [
                ms2_peak1_rt, ms2_peak2_rt, ms2_peak2_ratio, ms2_valley_frac, ms2_split_strength,
                ms3_peak1_rt, ms3_peak2_rt, ms3_peak2_ratio, ms3_valley_frac, ms3_split_strength,
            ]:
                lst.append(float("nan"))
            continue

        ms2_cons, ms2_per, ms2_var, ms2_n = analyze_skyline_for_peptide_variants(ms2_df, variants, cfg, ms2_peaks)
        ms3_cons, ms3_per, ms3_var, ms3_n = analyze_skyline_for_peptide_variants(ms3_df, variants, cfg, ms3_peaks)

        ms2_calls.append(ms2_cons)
        ms3_calls.append(ms3_cons)
        ms2_ntraces.append(ms2_n)
        ms3_ntraces.append(ms3_n)

        def strip_arrays(per_list: List[Dict]) -> List[Dict]:
            out = []
            for d in per_list:
                d2 = dict(d)
                d2.pop("rt", None)
                d2.pop("intensity", None)
                out.append(d2)
            return out

        ms2_detail.append(json.dumps(strip_arrays(ms2_per)))
        ms3_detail.append(json.dumps(strip_arrays(ms3_per)))

        chosen = ms2_var if ms2_n >= ms3_n else ms3_var
        if not chosen:
            chosen = variants[0]
        sky_norm_chosen.append(chosen)

        rep2 = pick_representative(ms2_per, ms2_cons)
        rep3 = pick_representative(ms3_per, ms3_cons)

        ms2_peak1_rt.append(rep2.get("peak1_rt", float("nan")))
        ms2_peak2_rt.append(rep2.get("peak2_rt", float("nan")))
        ms2_peak2_ratio.append(rep2.get("peak2_ratio", float("nan")))
        ms2_valley_frac.append(rep2.get("valley_frac", float("nan")))
        ms2_split_strength.append(rep2.get("split_strength", float("nan")))

        ms3_peak1_rt.append(rep3.get("peak1_rt", float("nan")))
        ms3_peak2_rt.append(rep3.get("peak2_rt", float("nan")))
        ms3_peak2_ratio.append(rep3.get("peak2_ratio", float("nan")))
        ms3_valley_frac.append(rep3.get("valley_frac", float("nan")))
        ms3_split_strength.append(rep3.get("split_strength", float("nan")))

        emit_traces(ms2_per, "MS2", gene, site, unip, match_cell)
        emit_traces(ms3_per, "MS3", gene, site, unip, match_cell)

    updated_unique = unique.copy()
    updated_unique["sky_norm_peptide_all"] = sky_norm_all
    updated_unique["sky_norm_peptide"] = sky_norm_chosen

    updated_unique["sky_split_call_MS2"] = ms2_calls
    updated_unique["sky_split_call_MS3"] = ms3_calls
    updated_unique["sky_MS2_n_precursor_traces"] = ms2_ntraces
    updated_unique["sky_MS3_n_precursor_traces"] = ms3_ntraces

    updated_unique["sky_MS2_peak1_rt"] = ms2_peak1_rt
    updated_unique["sky_MS2_peak2_rt"] = ms2_peak2_rt
    updated_unique["sky_MS2_peak2_ratio"] = ms2_peak2_ratio
    updated_unique["sky_MS2_valley_frac"] = ms2_valley_frac
    updated_unique["sky_MS2_split_strength"] = ms2_split_strength

    updated_unique["sky_MS3_peak1_rt"] = ms3_peak1_rt
    updated_unique["sky_MS3_peak2_rt"] = ms3_peak2_rt
    updated_unique["sky_MS3_peak2_ratio"] = ms3_peak2_ratio
    updated_unique["sky_MS3_valley_frac"] = ms3_valley_frac
    updated_unique["sky_MS3_split_strength"] = ms3_split_strength

    updated_unique["sky_MS2_per_precursor_details_json"] = ms2_detail
    updated_unique["sky_MS3_per_precursor_details_json"] = ms3_detail

    updated_unique["sky_split_call_combined_worst"] = [
        max([a, b], key=lambda c: SPLIT_ORDER.get(c, 0)) for a, b in zip(ms2_calls, ms3_calls)
    ]

    # outputs
    if output_xlsx_name:
        out_xlsx = out_dir / output_xlsx_name
    else:
        out_xlsx = out_dir / f"{inputs.sites_xlsx.stem}_with_skyline_splitting.xlsx"
    out_parquet = out_dir / traces_parquet_name

    _log(f"[splitting] Writing workbook: {out_xlsx}", log)
    propagate_to_sheets(inputs.sites_xlsx, out_xlsx, updated_unique, key_cols=key_cols)

    _log("[splitting] Adding Chromatograms sheet...", log)
    add_chromatograms_sheet(out_xlsx, chrom_rows_excel, "Chromatograms")

    _log(f"[splitting] Writing traces parquet: {out_parquet}", log)
    write_traces_parquet(trace_rows_parquet, out_parquet)

    _log("[splitting] Done.", log)
    return SplittingOutputs(
        output_xlsx=out_xlsx,
        traces_parquet=out_parquet,
        n_unique_sites=int(len(updated_unique)),
        n_chrom_rows=int(len(chrom_rows_excel)),
        n_trace_rows=int(len(trace_rows_parquet)),
    )


# -----------------------------
# Optional CLI wrapper
# -----------------------------

def main() -> None:
    base_dir = Path(__file__).resolve().parent

    inputs = SplittingInputs(
        sites_xlsx=base_dir / "merged_log2fc_results_sites.xlsx",
        sky_ms2_tsv=base_dir / "skyline_MS2_export.tsv",
        sky_ms3_tsv=base_dir / "skyline_MS3_export.tsv",
        ms2_peaks_csv=base_dir / "MS2_peaks_report.csv",
        ms3_peaks_csv=base_dir / "MS3_peaks_report.csv",
    )

    out = run(inputs, out_dir=base_dir, log=print)
    print(f"✅ Wrote: {out.output_xlsx}")
    print(f"✅ Wrote: {out.traces_parquet}")


if __name__ == "__main__":
    main()
